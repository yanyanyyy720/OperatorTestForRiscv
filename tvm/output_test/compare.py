import os
import struct
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

warnings.filterwarnings('ignore')


class AdvancedOperatorAnalyzer:
    """高级TVM算子统计分析工具（按配置分组）"""

    def __init__(self, op_lib_path):
        self.op_lib_path = op_lib_path
        self.results = []

    def load_multi_group_data(self, filename):
        """加载C++生成的二进制数据文件"""
        groups = []

        try:
            with open(filename, 'rb') as f:
                # 读取组数
                num_groups = struct.unpack('i', f.read(4))[0]

                for _ in range(num_groups):
                    # 读取当前组的张量数量
                    num_arrays = struct.unpack('i', f.read(4))[0]
                    group = []

                    for _ in range(num_arrays):
                        # 读取数据类型 (DLDataType结构: code, bits, lanes)
                        dtype_data = struct.unpack('BBH', f.read(4))
                        code, bits, lanes = dtype_data

                        # 映射到numpy数据类型
                        dtype_map = {
                            (0, 32): np.float32,  # kDLFloat, 32
                            (0, 16): np.float16,  # kDLFloat, 16
                            (4, 16): np.uint16,  # kDLBfloat, 16 (bfloat16)
                            (1, 32): np.int32,  # kDLInt, 32
                            (1, 64): np.int64,  # kDLInt, 64
                            (1, 8): np.int8,  # kDLInt, 8
                            (2, 8): np.uint8,  # kDLUInt, 8
                        }

                        np_dtype = dtype_map.get((code, bits), np.float32)

                        # 读取维度数
                        ndim = struct.unpack('i', f.read(4))[0]

                        # 读取形状
                        shape = struct.unpack('q' * ndim, f.read(8 * ndim))

                        # 计算元素总数和字节大小
                        num_elements = 1
                        for dim in shape:
                            num_elements *= dim

                        elem_size = (bits + 7) // 8
                        total_bytes = num_elements * elem_size

                        # 读取数据
                        data = f.read(total_bytes)

                        # 转换为numpy数组
                        if np_dtype == np.float16:
                            arr = np.frombuffer(data, dtype=np.uint16).astype(np.float16)
                        elif np_dtype == np.uint16:  # bfloat16
                            arr = np.frombuffer(data, dtype=np.uint16)
                        else:
                            arr = np.frombuffer(data, dtype=np_dtype)

                        # 重塑形状
                        arr = arr.reshape(shape)
                        group.append(arr)

                    groups.append(group)
        except Exception as e:
            print(f"加载文件 {filename} 时出错: {e}")
            return None

        return groups

    def read_config_file(self, config_path):
        """读取算子配置文件"""
        try:
            with open(config_path, 'r') as f:
                return f.read()
        except Exception as e:
            print(f"读取配置文件 {config_path} 时出错: {e}")
            return "配置读取失败"

    def cosine_similarity(self,a, b, epsilon=1e-10):
        """
        使用NumPy实现的余弦相似度计算函数，处理零向量情况
        """
        # 展平数组
        a_flat = a.flatten()
        b_flat = b.flatten()

        # 检查是否为零向量
        a_norm = np.linalg.norm(a_flat)
        b_norm = np.linalg.norm(b_flat)

        # 如果两个向量都是零向量，返回1.0（完全相似）
        if a_norm < epsilon and b_norm < epsilon:
            return 1.0

        # 如果只有一个向量是零向量，返回0.0（完全不相似）
        if a_norm < epsilon or b_norm < epsilon:
            return 0.0

        # 计算点积和模长
        dot_product = np.dot(a_flat, b_flat)
        norm_product = a_norm * b_norm

        # 计算余弦相似度
        cos_similarity = dot_product / norm_product

        # 确保结果在有效范围内
        cos_similarity = np.clip(cos_similarity, -1.0, 1.0)

        return cos_similarity

    def analyze_operator(self, set_name, op_name):
        """分析单个算子的数据"""
        data_path = os.path.join(self.op_lib_path, set_name, "data")
        config_path = os.path.join(self.op_lib_path, set_name, "native", f"{op_name}.cfg")

        # 读取配置文件
        config_content = self.read_config_file(config_path)

        # 加载输入和输出数据
        inputs = self.load_multi_group_data(os.path.join(data_path, f"{op_name}_inputs.bin"))
        ref_outputs = self.load_multi_group_data(os.path.join(data_path, f"{op_name}_ref_outputs.bin"))
        val_outputs = self.load_multi_group_data(os.path.join(data_path, f"{op_name}_val_outputs.bin"))

        if inputs is None or ref_outputs is None or val_outputs is None:
            return None

        # 基本统计信息
        num_groups = len(inputs)
        num_inputs_per_group = len(inputs[0]) if inputs else 0
        num_outputs_per_group = len(ref_outputs[0]) if ref_outputs else 0

        # 为每个组存储统计信息
        group_stats = []

        for group_idx, (ref_group, val_group) in enumerate(zip(ref_outputs, val_outputs)):
            if len(ref_group) != len(val_group):
                continue

            # 为当前组存储误差和相似度信息
            abs_errors = []
            rel_errors = []
            cos_similarities = []

            for output_idx, (ref_arr, val_arr) in enumerate(zip(ref_group, val_group)):
                if ref_arr.shape != val_arr.shape:
                    continue

                # 计算绝对误差和相对误差
                abs_error = np.abs(ref_arr - val_arr)
                rel_error = abs_error / (np.abs(ref_arr) + 1e-10)  # 避免除以零

                # 计算余弦相似度
                cos_sim = self.cosine_similarity(ref_arr, val_arr)

                # 展平误差数组以便统计
                abs_errors_flat = abs_error.flatten()
                rel_errors_flat = rel_error.flatten()
                ref_flat = ref_arr.flatten()

                abs_errors.extend(abs_errors_flat)
                rel_errors.extend(rel_errors_flat)
                cos_similarities.append(cos_sim)

            # 转换为numpy数组
            abs_errors = np.array(abs_errors)
            rel_errors = np.array(rel_errors)
            cos_similarities = np.array(cos_similarities)

            # 过滤无穷大和NaN值
            abs_errors = abs_errors[np.isfinite(abs_errors)]
            rel_errors = rel_errors[np.isfinite(rel_errors)]
            # cos_similarities = cos_similarities[np.isfinite(cos_similarities)]

            # 找出TopK绝对误差及其对应的真值
            k = min(5, len(abs_errors))
            if k > 0:
                topk_abs_indices = np.argpartition(abs_errors, -k)[-k:]
                topk_abs_errors = abs_errors[topk_abs_indices]

                # 找出TopK相对误差
                topk_rel_indices = np.argpartition(rel_errors, -k)[-k:]
                topk_rel_errors = rel_errors[topk_rel_indices]

            # 存储当前组的统计信息
            group_stats.append({
                'group_idx': group_idx,
                'abs_errors': abs_errors,
                'rel_errors': rel_errors,
                'cos_similarities': cos_similarities,
                'topk_abs_errors': topk_abs_errors if k > 0 else np.zeros(5),
                'topk_rel_errors': topk_rel_errors if k > 0 else np.zeros(5),
                'min_cos_sim': np.min(cos_similarities) if len(cos_similarities) > 0 else 0,
                'mean_cos_sim': np.mean(cos_similarities) if len(cos_similarities) > 0 else 0,
                'var_cos_sim': np.var(cos_similarities) if len(cos_similarities) > 0 else 0,
            })

        # 计算跨组的统计信息
        if not group_stats:
            return None

        # 提取所有组的TopK误差
        all_topk_abs_errors = np.array([gs['topk_abs_errors'] for gs in group_stats])
        all_topk_rel_errors = np.array([gs['topk_rel_errors'] for gs in group_stats])

        # 计算TopK误差的均值和方差（按位置）
        topk_abs_means = np.mean(all_topk_abs_errors, axis=0) if len(all_topk_abs_errors) > 0 else np.zeros(5)
        topk_abs_vars = np.var(all_topk_abs_errors, axis=0) if len(all_topk_abs_errors) > 0 else np.zeros(5)
        topk_rel_means = np.mean(all_topk_rel_errors, axis=0) if len(all_topk_rel_errors) > 0 else np.zeros(5)
        topk_rel_vars = np.var(all_topk_rel_errors, axis=0) if len(all_topk_rel_errors) > 0 else np.zeros(5)

        # 提取所有组的相似度统计
        all_min_cos_sim = np.array([gs['min_cos_sim'] for gs in group_stats])
        all_mean_cos_sim = np.array([gs['mean_cos_sim'] for gs in group_stats])
        all_var_cos_sim = np.array([gs['var_cos_sim'] for gs in group_stats])

        # 计算跨组的相似度统计
        min_cos_sim_min = np.min(all_min_cos_sim) if len(all_min_cos_sim) > 0 else 0
        min_cos_sim_mean = np.mean(all_min_cos_sim) if len(all_min_cos_sim) > 0 else 0
        min_cos_sim_var = np.var(all_min_cos_sim) if len(all_min_cos_sim) > 0 else 0

        mean_cos_sim_min = np.min(all_mean_cos_sim) if len(all_mean_cos_sim) > 0 else 0
        mean_cos_sim_mean = np.mean(all_mean_cos_sim) if len(all_mean_cos_sim) > 0 else 0
        mean_cos_sim_var = np.var(all_mean_cos_sim) if len(all_mean_cos_sim) > 0 else 0

        # 计算全局误差统计（所有组的所有误差）
        all_abs_errors = np.concatenate([gs['abs_errors'] for gs in group_stats])
        all_rel_errors = np.concatenate([gs['rel_errors'] for gs in group_stats])

        # 计算各种统计指标
        stats_dict = {
            'set_name': set_name,
            'op_name': op_name,
            'config_content': config_content,
            'num_groups': num_groups,
            'num_inputs_per_group': num_inputs_per_group,
            'num_outputs_per_group': num_outputs_per_group,
            'total_elements': len(all_abs_errors),

            # 全局误差统计
            'max_abs_error': np.max(all_abs_errors) if len(all_abs_errors) > 0 else 0,
            'max_rel_error': np.max(all_rel_errors) if len(all_rel_errors) > 0 else 0,
            'mean_abs_error': np.mean(all_abs_errors) if len(all_abs_errors) > 0 else 0,
            'median_abs_error': np.median(all_abs_errors) if len(all_abs_errors) > 0 else 0,
            'std_abs_error': np.std(all_abs_errors) if len(all_abs_errors) > 0 else 0,
            'var_abs_error': np.var(all_abs_errors) if len(all_abs_errors) > 0 else 0,
            'mean_rel_error': np.mean(all_rel_errors) if len(all_rel_errors) > 0 else 0,
            'median_rel_error': np.median(all_rel_errors) if len(all_rel_errors) > 0 else 0,
            'std_rel_error': np.std(all_rel_errors) if len(all_rel_errors) > 0 else 0,
            'var_rel_error': np.var(all_rel_errors) if len(all_rel_errors) > 0 else 0,
            'abs_error_95_percentile': np.percentile(all_abs_errors, 95) if len(all_abs_errors) > 0 else 0,
            'rel_error_95_percentile': np.percentile(all_rel_errors, 95) if len(all_rel_errors) > 0 else 0,

            # TopK绝对误差统计（跨组）
            'abs_top1_mean': topk_abs_means[0],
            'abs_top1_var': topk_abs_vars[0],
            'abs_top2_mean': topk_abs_means[1],
            'abs_top2_var': topk_abs_vars[1],
            'abs_top3_mean': topk_abs_means[2],
            'abs_top3_var': topk_abs_vars[2],
            'abs_top4_mean': topk_abs_means[3],
            'abs_top4_var': topk_abs_vars[3],
            'abs_top5_mean': topk_abs_means[4],
            'abs_top5_var': topk_abs_vars[4],

            # TopK相对误差统计（跨组）
            'rel_top1_mean': topk_rel_means[0],
            'rel_top1_var': topk_rel_vars[0],
            'rel_top2_mean': topk_rel_means[1],
            'rel_top2_var': topk_rel_vars[1],
            'rel_top3_mean': topk_rel_means[2],
            'rel_top3_var': topk_rel_vars[2],
            'rel_top4_mean': topk_rel_means[3],
            'rel_top4_var': topk_rel_vars[3],
            'rel_top5_mean': topk_rel_means[4],
            'rel_top5_var': topk_rel_vars[4],

            # 相似度统计（跨组）
            'min_cos_sim_min': min_cos_sim_min,
            'min_cos_sim_mean': min_cos_sim_mean,
            'min_cos_sim_var': min_cos_sim_var,
            'mean_cos_sim_min': mean_cos_sim_min,
            'mean_cos_sim_mean': mean_cos_sim_mean,
            'mean_cos_sim_var': mean_cos_sim_var,
        }

        return stats_dict

    def analyze_all_operators(self):
        """分析所有算子的数据"""
        # 获取所有set文件夹
        set_folders = [f for f in os.listdir(self.op_lib_path)
                       if os.path.isdir(os.path.join(self.op_lib_path, f)) and f.startswith('set_')]

        index = 0
        for set_folder in sorted(set_folders):
            # if index > 5:
            #     break
            index = index + 1
            set_path = os.path.join(self.op_lib_path, set_folder)
            data_path = os.path.join(set_path, "data")

            if not os.path.exists(data_path):
                continue

            # 获取所有算子名称（通过输入文件推断）
            input_files = [f for f in os.listdir(data_path) if f.endswith('_inputs.bin')]

            for input_file in input_files:
                op_name = input_file.replace('_inputs.bin', '')
                print(f"分析 {set_folder}/{op_name}...")

                stats_dict = self.analyze_operator(set_folder, op_name)
                if stats_dict:
                    self.results.append(stats_dict)

        return self.results

    def generate_excel_report(self, output_file="advanced_operator_analysis.xlsx"):
        """生成Excel报告"""
        if not self.results:
            print("没有分析结果可生成报告")
            return

        # 创建DataFrame
        df = pd.DataFrame(self.results)

        # 重新排列列的顺序，使相关统计量放在一起
        columns_order = [
            'set_name', 'op_name', 'config_content', 'num_groups',
            'num_inputs_per_group', 'num_outputs_per_group', 'total_elements',

            # 全局误差统计
            'max_abs_error', 'mean_abs_error', 'median_abs_error', 'std_abs_error', 'var_abs_error',
            'max_rel_error', 'mean_rel_error', 'median_rel_error', 'std_rel_error', 'var_rel_error',
            'abs_error_95_percentile', 'rel_error_95_percentile',

            # TopK绝对误差统计
            'abs_top1_mean', 'abs_top1_var',
            'abs_top2_mean', 'abs_top2_var',
            'abs_top3_mean', 'abs_top3_var',
            'abs_top4_mean', 'abs_top4_var',
            'abs_top5_mean', 'abs_top5_var',

            # TopK相对误差统计
            'rel_top1_mean', 'rel_top1_var',
            'rel_top2_mean', 'rel_top2_var',
            'rel_top3_mean', 'rel_top3_var',
            'rel_top4_mean', 'rel_top4_var',
            'rel_top5_mean', 'rel_top5_var',

            # 相似度统计
            'min_cos_sim_min', 'min_cos_sim_mean', 'min_cos_sim_var',
            'mean_cos_sim_min', 'mean_cos_sim_mean', 'mean_cos_sim_var',
        ]

        # 重新排列列
        df = df[columns_order]

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "算子分析汇总"

        # 设置标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 应用样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = {
            'A': 15, 'B': 20, 'C': 50, 'D': 10, 'E': 15, 'F': 15, 'G': 15,
            'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 15, 'N': 15,
            'O': 15, 'P': 15, 'Q': 15, 'R': 15, 'S': 15, 'T': 15, 'U': 15,
            'V': 15, 'W': 15, 'X': 15, 'Y': 15, 'Z': 15, 'AA': 15, 'AB': 15,
            'AC': 15, 'AD': 15, 'AE': 15, 'AF': 15, 'AG': 15, 'AH': 15, 'AI': 15,
            'AJ': 15, 'AK': 15, 'AL': 15, 'AM': 15, 'AN': 15, 'AO': 15, 'AP': 15,
            'AQ': 15, 'AR': 15, 'AS': 15, 'AT': 15, 'AU': 15, 'AV': 15, 'AW': 15,
            'AX': 15, 'AY': 15, 'AZ': 15, 'BA': 15, 'BB': 15, 'BC': 15, 'BD': 15,
            'BE': 15, 'BF': 15, 'BG': 15, 'BH': 15, 'BI': 15, 'BJ': 15, 'BK': 15,
            'BL': 15, 'BM': 15, 'BN': 15, 'BO': 15, 'BP': 15, 'BQ': 15, 'BR': 15
        }

        for col, width in column_widths.items():
            if col in ws.column_dimensions:
                ws.column_dimensions[col].width = width

        # 添加条件格式
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # 对误差较大的单元格标记颜色
        for row in range(2, len(df) + 2):
            # 绝对误差大于1e-5标记红色
            max_abs_error = ws[f'H{row}'].value
            if max_abs_error is not None and max_abs_error > 1e-5:
                ws[f'H{row}'].fill = red_fill

            # 相对误差大于0.01标记红色
            max_rel_error = ws[f'M{row}'].value
            if max_rel_error is not None and max_rel_error > 0.01:
                ws[f'M{row}'].fill = red_fill

            # 最小余弦相似度小于0.99标记黄色
            min_cos_sim_min = ws[f'AQ{row}'].value
            if min_cos_sim_min is not None and min_cos_sim_min < 0.99:
                ws[f'AQ{row}'].fill = yellow_fill

            # 最小余弦相似度大于0.999标记绿色
            if min_cos_sim_min is not None and min_cos_sim_min > 0.999:
                ws[f'AQ{row}'].fill = green_fill

        # 保存文件
        wb.save(output_file)
        print(f"报告已保存到 {output_file}")

        return df


def main():
    """主函数"""
    # 设置op_lib文件夹路径
    op_lib_path = "op_lib_2/op_lib"  # 请根据实际情况修改路径

    # 创建分析器
    analyzer = AdvancedOperatorAnalyzer(op_lib_path)

    # 分析所有算子
    print("开始分析算子数据...")
    results = analyzer.analyze_all_operators()

    if not results:
        print("未找到任何算子数据")
        return

    # 生成Excel报告
    print("生成Excel报告...")
    df = analyzer.generate_excel_report()

    # 打印简要统计
    print(f"\n分析完成！共分析 {len(df)} 个算子")
    print(f"最大绝对误差: {df['max_abs_error'].max():.6e}")
    print(f"最大相对误差: {df['max_rel_error'].max():.6e}")
    print(f"最小余弦相似度最小值: {df['min_cos_sim_min'].min():.6f}")


if __name__ == "__main__":
    main()